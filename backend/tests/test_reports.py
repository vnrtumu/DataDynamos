"""Benchmark report endpoint tests — mock engines, fully offline."""

from io import BytesIO

from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlmodel import Session

from app.db import engine
from app.main import app
from app.models import Document, PipelineRun

from .conftest import SAMPLES


def _upload(client: TestClient, name: str) -> str:
    with (SAMPLES / name).open("rb") as fh:
        resp = client.post("/documents", files={"file": (name, fh)})
    assert resp.status_code == 201, resp.text
    return resp.json()["id"]


def _seed_doc(page_count: int = 2) -> str:
    """Insert a document + a fully-decided run with known latencies and costs."""
    with Session(engine) as session:
        doc = Document(
            filename="bench.pdf",
            mime="application/pdf",
            doc_type="cms1500",
            page_count=page_count,
            status="decided",
        )
        session.add(doc)
        session.commit()
        session.refresh(doc)
        run = PipelineRun(
            document_id=doc.id,
            stage_results={
                "ocr": {
                    "docling": {
                        "engine_name": "docling",
                        "device": "cpu",
                        "latency_ms": 800,
                        "avg_confidence": 0.94,
                        "pages": [],
                    }
                },
                "structure": {
                    "doc_type": "cms1500",
                    "ocr_engine": "docling",
                    "latency_ms": 150,
                    "extraction_confidence": 0.95,
                    "cost_summary": {
                        "tier": "Tier A",
                        "preprocessing_cost": 0.0002,
                        "ocr_engine_cost": 0.0010,
                        "vlm_llm_cost": 0.0,
                        "total_cost": 0.0012,
                    },
                    "accuracy_metrics": {
                        "overall_accuracy": 96.0,
                        "field_accuracy": 97.0,
                        "rule_pass_rate": 100.0,
                        "ocr_confidence": 94.0,
                        "grounded_ratio": 91.0,
                    },
                },
                "decide": {"latency_ms": 50, "decision": "approve"},
            },
        )
        session.add(run)
        session.commit()
        return doc.id


def test_benchmark_report_end_to_end():
    """Mock pipeline via the API produces a valid workbook with 3 sheets."""
    with TestClient(app) as client:
        assert client.delete("/documents").status_code == 204
        _seed_doc()
        resp = client.get("/reports/benchmark.xlsx")
        assert resp.status_code == 200, resp.text
        assert "spreadsheetml" in resp.headers["content-type"]
        assert "benchmark_report.xlsx" in resp.headers["content-disposition"]

        wb = load_workbook(BytesIO(resp.content))
        assert wb.sheetnames == ["Overall Metrics", "Per Document", "Cost Analysis"]

        per_doc = wb["Per Document"]
        assert per_doc.max_row == 2  # header + 1 seeded doc
        assert per_doc.cell(row=2, column=3).value == 2

        cost = wb["Cost Analysis"]
        headers = [cost.cell(row=1, column=c).value for c in range(1, 10)]
        assert headers == [
            "Document", "Pages", "OCR ($/pg)", "LLM ($/pg)", "Vision AI ($/pg)",
            "GPU ($/pg)", "CPU ($/pg)", "Total Cost per Page ($)", "Total Cost ($)",
        ]
        total_col = cost.cell(row=2, column=9).value
        assert cost.cell(row=3, column=9).value == total_col  # aggregate row


def test_benchmark_report_metrics_math():
    """Seeded run with known latencies/costs yields exact overall metrics."""
    with TestClient(app) as client:
        assert client.delete("/documents").status_code == 204
        _seed_doc(page_count=2)
        resp = client.get("/reports/benchmark.xlsx")
        wb = load_workbook(BytesIO(resp.content))

        overview = wb["Overall Metrics"]
        metrics = {
            overview.cell(row=r, column=1).value: overview.cell(row=r, column=2).value
            for r in range(3, 12)
        }
        # 800ms OCR + 150ms structure + 50ms decide = 1000ms across 2 pages.
        assert metrics["Total Pages Processed"] == 2
        assert metrics["Total Documents"] == 1
        assert metrics["Processing Time (s)"] == 1.0
        assert metrics["Average Latency (ms/page)"] == 500.0
        assert metrics["Pages per Second"] == 2.0
        assert metrics["Accuracy (%)"] == 96.0
        assert metrics["Precision (%)"] == 97.0
        assert metrics["Recall (%)"] == 91.0

        cost = wb["Cost Analysis"]
        # docling @ 0.0005/pg: OCR fee = 0.0005; CPU = 0.0001 + 0.0005 (cpu device);
        # LLM = 0.0015 / 2; Vision = 0; Total/pg = CPU + OCR + LLM.
        cpu, ocr, llm = cost.cell(row=2, column=7).value, cost.cell(row=2, column=3).value, cost.cell(row=2, column=4).value
        total_per_page = cost.cell(row=2, column=8).value
        assert cpu == 0.0006
        assert ocr == 0.0005
        assert llm == 0.00075
        assert cost.cell(row=2, column=6).value == 0  # GPU: 0 on cpu device
        assert total_per_page == round(cpu + ocr + llm, 6)
        assert cost.cell(row=2, column=9).value == round(total_per_page * 2, 6)


def test_benchmark_report_empty_dataset():
    with TestClient(app) as client:
        assert client.delete("/documents").status_code == 204
        resp = client.get("/reports/benchmark.xlsx")
        assert resp.status_code == 200, resp.text
        wb = load_workbook(BytesIO(resp.content))
        assert wb["Overall Metrics"].cell(row=3, column=2).value == 0
        assert wb["Per Document"].max_row == 1
